"""
Microsoft 365 Audit Log Analyzer — Per-User PDF Report Generator
Run from the repo root:  python generate_reports.py
"""

import pandas as pd
import numpy as np
import os
import glob
import json
import re
import shutil
from datetime import datetime, timedelta
from collections import defaultdict

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
import seaborn as sns
from fpdf import FPDF

sns.set_style("whitegrid")
plt.rcParams.update({"figure.max_open_warning": 0})

# ──────────────────────────────────────────────
# CONFIG
# ──────────────────────────────────────────────
INPUT_DIR  = "input"
OUTPUT_DIR = "output"
TEMP_DIR   = os.path.join(OUTPUT_DIR, ".tmp_charts")
RUN_DATE   = datetime.now().strftime("%Y-%m-%d")

# ──────────────────────────────────────────────
# LOAD
# ──────────────────────────────────────────────
csv_files = glob.glob(os.path.join(INPUT_DIR, "*.csv"))
print(f"Found {len(csv_files)} CSV file(s)")

frames = []
for f in csv_files:
    df = pd.read_csv(f, encoding="utf-8-sig")
    print(f"  {os.path.basename(f)}: {len(df)} rows")
    frames.append(df)

raw_df = pd.concat(frames, ignore_index=True)
print(f"Combined: {raw_df.shape[0]} rows x {raw_df.shape[1]} columns")

# ──────────────────────────────────────────────
# PARSE & CLEAN
# ──────────────────────────────────────────────
raw_df["CreationDate"] = pd.to_datetime(raw_df["CreationDate"].str.strip(), utc=True)

def safe_json(s):
    try:
        return json.loads(s)
    except Exception:
        return {}

print("Parsing AuditData JSON (may take a moment)...")
audit_parsed = raw_df["AuditData"].apply(safe_json)
for field in ["SiteUrl", "Workload", "SourceFileName", "SourceFileExtension",
              "SourceRelativeUrl", "ClientIP", "ItemType", "ObjectId",
              "Platform", "GeoLocation"]:
    raw_df[field] = audit_parsed.apply(lambda d, f=field: d.get(f, ""))

OP_CATEGORY = {
    "FileDeleted": "Delete", "FileRecycled": "Delete", "FolderRecycled": "Delete",
    "ListItemDeleted": "Delete",
    "FileDownloaded": "Download (Manual)", "FileSyncDownloadedFull": "Sync (Auto)",
    "FileAccessed": "Read / Access", "FileAccessedExtended": "Read / Access",
    "FilePreviewed": "Read / Access", "PageViewed": "Read / Access",
    "PageViewedExtended": "Read / Access", "ListViewed": "Read / Access",
    "AttachmentAccess": "Read / Access",
    "FileModified": "Modify", "FileModifiedExtended": "Modify", "FolderModified": "Modify",
    "FileUploaded": "Upload", "FileUploadedPartial": "Upload", "FileSyncUploadedFull": "Upload",
    "FileRenamed": "Rename / Move", "FileMoved": "Rename / Move",
    "FolderMoved": "Rename / Move", "FolderRenamed": "Rename / Move",
    "FolderCreated": "Create Folder",
    "SharingSet": "Share", "SharingInheritanceBroken": "Share",
    "SharingLinkCreated": "Share", "SharingLinkUsed": "Share",
    "SharingLinkUpdated": "Share", "AddedToSharingLink": "Share",
    "SecureLinkCreated": "Share", "SecureLinkUsed": "Share",
    "SecureLinkUpdated": "Share", "AddedToSecureLink": "Share",
    "CompanyLinkUsed": "Share",
    "AddedToGroup": "Permission Change",
    "UserLoggedIn": "Login", "UserLoginFailed": "Login Failed", "SignInEvent": "Login",
    "MessageSent": "Teams / Comms", "MessageCreatedHasLink": "Teams / Comms",
    "MessageUpdated": "Teams / Comms", "MessageEditedHasLink": "Teams / Comms",
    "ReactedToMessage": "Teams / Comms", "TeamsSessionStarted": "Teams / Comms",
    "MeetingParticipantDetail": "Teams / Comms", "MeetingDetail": "Teams / Comms",
    "CallParticipantDetail": "Teams / Comms",
    "TeamsMeetingRecordingUploaded": "Teams / Comms",
    "MailItemsAccessed": "Email", "Send": "Email",
    "TaskModified": "Planner", "TaskCompleted": "Planner",
    "TaskListRead": "Planner", "PlanRead": "Planner",
    "ShortcutAdded": "Other",
    "BaselineSecurityModeThirdPartyAppHPA": "Other",
}

raw_df["OpCategory"] = raw_df["Operation"].map(OP_CATEGORY).fillna("Other")

def clean_site(url):
    if not url or not isinstance(url, str) or not url.startswith("http"):
        return url if isinstance(url, str) and url else "N/A"
    m = re.search(r"/sites/([^/]+)", url)
    return m.group(1) if m else re.sub(r"https?://[^/]+", "", url).strip("/") or url

raw_df["SiteName"] = raw_df["SiteUrl"].apply(clean_site)
raw_df["Date"]     = raw_df["CreationDate"].dt.date
raw_df["Hour"]     = raw_df["CreationDate"].dt.hour

print(f"Rows after cleaning: {len(raw_df)}")
print(f"Operation categories: {sorted(raw_df['OpCategory'].unique())}")
print(f"Date range: {raw_df['Date'].min()} to {raw_df['Date'].max()}")

users = sorted(raw_df["UserId"].unique())
print(f"Users: {users}")

# ──────────────────────────────────────────────
# EXPORT HUMAN-READABLE EXCEL FILES
# ──────────────────────────────────────────────
os.makedirs(OUTPUT_DIR, exist_ok=True)

# Select and rename columns for human-readable output
EXPORT_COLUMNS = {
    "CreationDate":        "Timestamp (UTC)",
    "UserId":              "User",
    "Operation":           "Raw Operation",
    "OpCategory":          "Category",
    "Workload":            "Service",
    "SiteName":            "Site",
    "SiteUrl":             "Site URL",
    "SourceFileName":      "File Name",
    "SourceFileExtension": "File Extension",
    "SourceRelativeUrl":   "Folder Path",
    "ItemType":            "Item Type",
    "ClientIP":            "Client IP",
    "Platform":            "Platform",
    "GeoLocation":         "Geolocation",
    "ObjectId":            "Full Resource Path",
}

from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from openpyxl.styles import numbers as xl_numbers

def _autofit_columns(ws, df):
    """Auto-fit column widths based on content."""
    for col_idx, col_name in enumerate(df.columns, 1):
        max_len = max(len(str(col_name)),
                      df[col_name].astype(str).str.len().max() if len(df) > 0 else 0)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)

def _add_named_table(ws, name, n_rows, n_cols):
    """Add a proper Excel Table with banded rows."""
    ref = f"A1:{get_column_letter(n_cols)}{n_rows + 1}"
    tbl = Table(displayName=name, ref=ref)
    tbl.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9", showFirstColumn=False,
        showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    ws.add_table(tbl)

excel_generated = []
for user in users:
    udf = raw_df[raw_df["UserId"] == user].copy()
    # Build readable dataframe — only columns that exist
    avail = {k: v for k, v in EXPORT_COLUMNS.items() if k in udf.columns}
    export_df = udf[list(avail.keys())].rename(columns=avail)
    # Keep Timestamp as real datetime (tz-naive for Excel compatibility)
    export_df["Timestamp (UTC)"] = export_df["Timestamp (UTC)"].dt.tz_localize(None)
    export_df = export_df.sort_values("Timestamp (UTC)")
    # Add a Date-only column for easier pivot filtering
    export_df.insert(1, "Date", export_df["Timestamp (UTC)"].dt.date)

    safe_name = re.sub(r'[^\w.@-]', '_', user)
    xlsx_path = os.path.join(OUTPUT_DIR, f"audit_records_{safe_name}_{RUN_DATE}.xlsx")

    with pd.ExcelWriter(xlsx_path, engine="openpyxl",
                         datetime_format="YYYY-MM-DD HH:MM:SS",
                         date_format="YYYY-MM-DD") as writer:

        # --- Sheet 1: Audit Records (all rows, proper Table) ---
        export_df.to_excel(writer, index=False, sheet_name="Audit Records")
        ws_data = writer.sheets["Audit Records"]
        _autofit_columns(ws_data, export_df)
        ws_data.freeze_panes = "A2"
        # Sanitise table name (Excel requires letters/underscores only)
        tbl_name = "AuditRecords_" + re.sub(r'[^A-Za-z0-9]', '_', user)
        _add_named_table(ws_data, tbl_name, len(export_df), len(export_df.columns))
        # Format datetime column A with proper number format
        for row in range(2, len(export_df) + 2):
            ws_data.cell(row, 1).number_format = "YYYY-MM-DD HH:MM:SS"
        # Format date column B
        for row in range(2, len(export_df) + 2):
            ws_data.cell(row, 2).number_format = "YYYY-MM-DD"

        # --- Sheet 2: Summary by Date & Category (pivot-style) ---
        pivot_date = (export_df.groupby(["Date", "Category"])
                      .size().reset_index(name="Count"))
        pivot_date_wide = pivot_date.pivot_table(
            index="Date", columns="Category", values="Count",
            aggfunc="sum", fill_value=0)
        pivot_date_wide["Total"] = pivot_date_wide.sum(axis=1)
        pivot_date_wide = pivot_date_wide.sort_index()
        pivot_date_wide.index = pd.to_datetime(pivot_date_wide.index)
        # Write
        pivot_date_wide.to_excel(writer, sheet_name="By Date")
        ws_date = writer.sheets["By Date"]
        _autofit_columns(ws_date, pivot_date_wide.reset_index())
        ws_date.freeze_panes = "B2"
        tbl_name2 = "ByDate_" + re.sub(r'[^A-Za-z0-9]', '_', user)
        _add_named_table(ws_date, tbl_name2,
                         len(pivot_date_wide), len(pivot_date_wide.columns) + 1)
        # Format date column A
        for row in range(2, len(pivot_date_wide) + 2):
            ws_date.cell(row, 1).number_format = "YYYY-MM-DD"

        # --- Sheet 3: Summary by Site & Category (pivot-style) ---
        pivot_site = (export_df.groupby(["Site", "Category"])
                      .size().reset_index(name="Count"))
        pivot_site_wide = pivot_site.pivot_table(
            index="Site", columns="Category", values="Count",
            aggfunc="sum", fill_value=0)
        pivot_site_wide["Total"] = pivot_site_wide.sum(axis=1)
        pivot_site_wide = pivot_site_wide.sort_values("Total", ascending=False)
        # Write
        pivot_site_wide.to_excel(writer, sheet_name="By Site")
        ws_site = writer.sheets["By Site"]
        _autofit_columns(ws_site, pivot_site_wide.reset_index())
        ws_site.freeze_panes = "B2"
        tbl_name3 = "BySite_" + re.sub(r'[^A-Za-z0-9]', '_', user)
        _add_named_table(ws_site, tbl_name3,
                         len(pivot_site_wide), len(pivot_site_wide.columns) + 1)

    size = os.path.getsize(xlsx_path)
    excel_generated.append((user, xlsx_path, len(export_df), size))
    print(f"Excel: {xlsx_path}  ({len(export_df)} rows, {size:,} bytes)")
    print(f"  Sheets: Audit Records, By Date ({len(pivot_date_wide)} dates), "
          f"By Site ({len(pivot_site_wide)} sites)")

print(f"\nExported {len(excel_generated)} user Excel file(s).")

# ──────────────────────────────────────────────
# ANALYSIS HELPERS
# ──────────────────────────────────────────────
SENSITIVE_CATEGORIES = {"Delete", "Download (Manual)", "Share"}

def compute_op_counts(udf):
    counts = udf["OpCategory"].value_counts().rename("Count").to_frame()
    counts["Pct"] = (counts["Count"] / counts["Count"].sum() * 100).round(1)
    return counts.sort_values("Count", ascending=False)

def compute_site_cross(udf, top_n=12):
    ct = pd.crosstab(udf["SiteName"], udf["OpCategory"])
    ct["_Total"] = ct.sum(axis=1)
    ct = ct.sort_values("_Total", ascending=False).head(top_n)
    ct = ct.drop(columns=["_Total"])
    return ct

def compute_daily_timeline(udf):
    return udf.groupby("Date").size()

def detect_risk_indicators(udf, daily):
    indicators = []
    total = len(udf)
    date_min, date_max = udf["Date"].min(), udf["Date"].max()
    period_days = max((date_max - date_min).days, 1)

    n_del = len(udf[udf["OpCategory"] == "Delete"])
    if n_del > 50:
        indicators.append(("HIGH", "Bulk deletions",
            f"{n_del} delete operations detected ({n_del/total*100:.1f}% of total)."))
    elif n_del > 10:
        indicators.append(("MEDIUM", "Moderate deletions",
            f"{n_del} delete operations detected."))

    n_dl_manual = len(udf[udf["OpCategory"] == "Download (Manual)"])
    n_sync = len(udf[udf["OpCategory"] == "Sync (Auto)"])
    if n_dl_manual > 50:
        indicators.append(("HIGH", "Mass manual downloads",
            f"{n_dl_manual} manual download ops. Potential data exfiltration."))
    elif n_dl_manual > 20:
        indicators.append(("MEDIUM", "Notable manual downloads",
            f"{n_dl_manual} manual download ops."))
    # Flag sync only if disproportionate
    sync_daily = n_sync / period_days
    if sync_daily > 500:
        indicators.append(("MEDIUM", "Very high sync volume",
            f"{n_sync} auto-sync ops ({sync_daily:.0f}/day). May indicate large-scale data pull via OneDrive."))

    n_share = len(udf[udf["OpCategory"] == "Share"])
    if n_share > 30:
        indicators.append(("HIGH", "Excessive sharing",
            f"{n_share} sharing operations. Review shared targets."))
    elif n_share > 10:
        indicators.append(("MEDIUM", "Notable sharing activity",
            f"{n_share} sharing operations."))

    off = udf[(udf["Hour"] < 7) | (udf["Hour"] >= 20)]
    pct_off = len(off) / total * 100 if total else 0
    if pct_off > 15:
        indicators.append(("MEDIUM", "Significant off-hours activity",
            f"{len(off)} ops ({pct_off:.1f}%) outside 07:00-20:00."))
    elif pct_off > 5:
        indicators.append(("LOW", "Some off-hours activity",
            f"{len(off)} ops ({pct_off:.1f}%) outside 07:00-20:00."))

    if period_days >= 7:
        last_week_start = date_max - timedelta(days=6)
        lw = udf[udf["Date"] >= last_week_start]
        earlier = udf[udf["Date"] < last_week_start]
        if len(earlier) > 0:
            earlier_daily = len(earlier) / max((last_week_start - date_min).days, 1)
            lw_daily = len(lw) / 7
            if earlier_daily > 0 and lw_daily / earlier_daily > 2:
                indicators.append(("HIGH", "Activity spike in final week",
                    f"Last 7 days avg {lw_daily:.0f} ops/day vs prior {earlier_daily:.0f}/day "
                    f"({lw_daily/earlier_daily:.1f}x increase)."))
            elif earlier_daily > 0 and lw_daily / earlier_daily > 1.4:
                indicators.append(("MEDIUM", "Elevated last-week activity",
                    f"Last 7 days avg {lw_daily:.0f} ops/day vs prior {earlier_daily:.0f}/day."))
            # Detect sudden drop-off (user may have already left)
            if earlier_daily > 10 and lw_daily < earlier_daily * 0.2:
                indicators.append(("LOW", "Sudden activity drop in final week",
                    f"Last 7 days avg {lw_daily:.0f} ops/day vs prior {earlier_daily:.0f}/day. "
                    "User may have already departed — review prior period for exfiltration."))

    if len(daily) > 3:
        avg = daily.mean()
        spikes = daily[daily > 3 * avg]
        if len(spikes) > 0:
            spike_dates = ", ".join(str(d) for d in spikes.index[:5])
            indicators.append(("MEDIUM", f"{len(spikes)} spike day(s)",
                f"Days with >3x average volume: {spike_dates}"))

    n_fail = len(udf[udf["OpCategory"] == "Login Failed"])
    if n_fail >= 5:
        indicators.append(("MEDIUM", "Multiple login failures",
            f"{n_fail} failed login attempts detected."))

    if not indicators:
        indicators.append(("LOW", "No significant anomalies",
            "Activity appears within normal parameters."))
    return indicators

def compute_risk_score(indicators):
    score = 0
    for sev, _, _ in indicators:
        if sev == "HIGH":   score += 30
        elif sev == "MEDIUM": score += 15
        elif sev == "LOW":   score += 5
    return min(score, 100)

def top_files_for_sensitive_ops(udf, n=15):
    sens = udf[udf["OpCategory"].isin(SENSITIVE_CATEGORIES) & (udf["SourceFileName"] != "")]
    if sens.empty:
        return pd.DataFrame(columns=["File", "Operation", "Count"])
    top = (sens.groupby(["SourceFileName", "OpCategory"])
           .size().reset_index(name="Count")
           .sort_values("Count", ascending=False)
           .head(n))
    return top.rename(columns={"SourceFileName": "File", "OpCategory": "Operation"})

# ──────────────────────────────────────────────
# CHART HELPERS
# ──────────────────────────────────────────────
def chart_op_counts(counts_df, user, path):
    fig, ax = plt.subplots(figsize=(7, max(3, len(counts_df) * 0.4)))
    colors = []
    for cat in counts_df.index:
        if cat == "Delete":           colors.append("#e74c3c")
        elif cat == "Download / Sync": colors.append("#e67e22")
        elif cat == "Share":          colors.append("#f1c40f")
        else:                         colors.append("#3498db")
    ax.barh(counts_df.index[::-1], counts_df["Count"].values[::-1], color=colors[::-1])
    ax.set_xlabel("Count")
    ax.set_title(f"Operations by Category - {user}")
    for i, (v, p) in enumerate(zip(counts_df["Count"].values[::-1], counts_df["Pct"].values[::-1])):
        ax.text(v + max(counts_df["Count"]) * 0.01, i, f" {v}  ({p}%)", va="center", fontsize=8)
    plt.tight_layout()
    fig.savefig(path, dpi=130)
    plt.close(fig)

def chart_daily_timeline(daily, user, path, spike_threshold=None):
    fig, ax = plt.subplots(figsize=(8, 3))
    dates = list(daily.index)
    ax.bar(dates, daily.values, color="#3498db", width=0.8)
    if spike_threshold is not None:
        ax.axhline(spike_threshold, color="#e74c3c", ls="--", lw=1,
                   label=f"3x avg ({spike_threshold:.0f})")
        ax.legend(fontsize=8)
    ax.set_ylabel("Operations")
    ax.set_title(f"Daily Activity Timeline - {user}")
    fig.autofmt_xdate(rotation=45)
    plt.tight_layout()
    fig.savefig(path, dpi=130)
    plt.close(fig)

def chart_hourly(udf, user, path):
    fig, ax = plt.subplots(figsize=(7, 3))
    hour_counts = udf.groupby("Hour").size().reindex(range(24), fill_value=0)
    colors = ["#e67e22" if h < 7 or h >= 20 else "#3498db" for h in range(24)]
    ax.bar(range(24), hour_counts.values, color=colors)
    ax.set_xlabel("Hour of Day (UTC)")
    ax.set_ylabel("Operations")
    ax.set_title(f"Activity by Hour - {user}")
    ax.set_xticks(range(0, 24, 2))
    plt.tight_layout()
    fig.savefig(path, dpi=130)
    plt.close(fig)

def chart_site_heatmap(site_cross, user, path):
    if site_cross.empty:
        return False
    fig, ax = plt.subplots(figsize=(max(6, len(site_cross.columns) * 0.9),
                                     max(3, len(site_cross) * 0.5)))
    sns.heatmap(site_cross, annot=True, fmt="d", cmap="YlOrRd", ax=ax,
                linewidths=0.5, cbar_kws={"label": "Count"})
    ax.set_title(f"Operations by Site - {user}")
    ax.set_ylabel("")
    plt.tight_layout()
    fig.savefig(path, dpi=130)
    plt.close(fig)
    return True

# ──────────────────────────────────────────────
# PDF CLASS
# ──────────────────────────────────────────────
class AuditPDF(FPDF):
    def __init__(self, user, date_range, risk_score):
        super().__init__()
        self.user = user
        self.date_range = date_range
        self.risk_score = risk_score

    def header(self):
        self.set_font("Helvetica", "B", 14)
        self.cell(0, 8, "Microsoft 365 Audit - User Activity Report",
                  new_x="LMARGIN", new_y="NEXT", align="C")
        self.set_font("Helvetica", "", 9)
        self.cell(0, 5,
                  f"User: {self.user}    |    Period: {self.date_range}    |    "
                  f"Generated: {datetime.now():%Y-%m-%d %H:%M}",
                  new_x="LMARGIN", new_y="NEXT", align="C")
        self.ln(2)
        self.set_draw_color(200, 200, 200)
        self.line(10, self.get_y(), self.w - 10, self.get_y())
        self.ln(3)

    def footer(self):
        self.set_y(-15)
        self.set_font("Helvetica", "I", 8)
        self.set_text_color(128)
        self.cell(0, 10, f"Page {self.page_no()}/{{nb}}", align="C")
        self.set_text_color(0)

    def section_title(self, title):
        self.set_font("Helvetica", "B", 12)
        self.set_fill_color(41, 128, 185)
        self.set_text_color(255)
        self.cell(0, 8, f"  {title}", new_x="LMARGIN", new_y="NEXT", fill=True)
        self.set_text_color(0)
        self.ln(3)

    def risk_badge(self, score):
        self.set_font("Helvetica", "B", 11)
        if score >= 60:
            self.set_fill_color(231, 76, 60); label = "HIGH RISK"
        elif score >= 30:
            self.set_fill_color(243, 156, 18); label = "MEDIUM RISK"
        else:
            self.set_fill_color(46, 204, 113); label = "LOW RISK"
        self.set_text_color(255)
        self.cell(50, 10, f"  Risk Score: {score}/100 - {label}  ",
                  new_x="LMARGIN", new_y="NEXT", fill=True)
        self.set_text_color(0)
        self.ln(4)

    def safe_text(self, text):
        """Replace unicode chars that Helvetica can't render."""
        return str(text).replace("\u2014", "-").replace("\u2013", "-").replace("\u2018", "'").replace("\u2019", "'").replace("\u201c", '"').replace("\u201d", '"').replace("\u2026", "...")

    def add_table(self, headers, rows, col_widths=None):
        if col_widths is None:
            col_widths = [(self.w - 20) / len(headers)] * len(headers)
        self.set_font("Helvetica", "B", 8)
        self.set_fill_color(220, 220, 220)
        for i, h in enumerate(headers):
            self.cell(col_widths[i], 6, self.safe_text(h), border=1, fill=True, align="C")
        self.ln()
        self.set_font("Helvetica", "", 8)
        for row in rows:
            if self.get_y() > self.h - 30:
                self.add_page()
            for i, val in enumerate(row):
                self.cell(col_widths[i], 5, self.safe_text(str(val)[:60]), border=1, align="C")
            self.ln()

    def safe_cell(self, w, h, txt, **kwargs):
        self.cell(w, h, self.safe_text(txt), **kwargs)

    def safe_multi_cell(self, w, h, txt, **kwargs):
        self.multi_cell(w, h, self.safe_text(txt), **kwargs)

    def add_image_safe(self, path, w=170):
        if self.get_y() > self.h - 90:
            self.add_page()
        self.image(path, x=15, w=w)
        self.ln(5)


# ──────────────────────────────────────────────
# GENERATE
# ──────────────────────────────────────────────
os.makedirs(OUTPUT_DIR, exist_ok=True)
os.makedirs(TEMP_DIR, exist_ok=True)

generated = []
for user in users:
    print(f"\n{'='*60}")
    print(f"Generating report for: {user}")
    print(f"{'='*60}")

    udf = raw_df[raw_df["UserId"] == user].copy()
    date_min, date_max = udf["Date"].min(), udf["Date"].max()
    date_range = f"{date_min} to {date_max}"

    op_counts   = compute_op_counts(udf)
    site_cross  = compute_site_cross(udf)
    daily       = compute_daily_timeline(udf)
    indicators  = detect_risk_indicators(udf, daily)
    risk_score  = compute_risk_score(indicators)
    top_files   = top_files_for_sensitive_ops(udf)

    last_week_start = date_max - timedelta(days=6)
    lw_df    = udf[udf["Date"] >= last_week_start]
    prior_df = udf[udf["Date"] < last_week_start]

    # Charts
    p_ops   = os.path.join(TEMP_DIR, f"{re.sub(r'[^a-zA-Z0-9]', '_', user)}_ops.png")
    p_daily = os.path.join(TEMP_DIR, f"{re.sub(r'[^a-zA-Z0-9]', '_', user)}_daily.png")
    p_hour  = os.path.join(TEMP_DIR, f"{re.sub(r'[^a-zA-Z0-9]', '_', user)}_hour.png")
    p_site  = os.path.join(TEMP_DIR, f"{re.sub(r'[^a-zA-Z0-9]', '_', user)}_site.png")

    chart_op_counts(op_counts, user, p_ops)
    spike_thr = daily.mean() * 3 if len(daily) > 3 else None
    chart_daily_timeline(daily, user, p_daily, spike_thr)
    chart_hourly(udf, user, p_hour)
    has_site_chart = chart_site_heatmap(site_cross, user, p_site)

    # Build PDF
    pdf = AuditPDF(user, date_range, risk_score)
    pdf.alias_nb_pages()
    pdf.set_auto_page_break(auto=True, margin=20)
    pdf.add_page()

    # 1. Executive Summary
    pdf.section_title("Executive Summary")
    pdf.risk_badge(risk_score)
    pdf.set_font("Helvetica", "", 9)
    pdf.cell(0, 5,
             f"Total operations: {len(udf):,}        Period: {date_range}        "
             f"Days active: {udf['Date'].nunique()} / {(date_max - date_min).days + 1}",
             new_x="LMARGIN", new_y="NEXT")
    pdf.ln(2)

    pdf.set_font("Helvetica", "B", 9)
    pdf.cell(0, 5, "Risk Indicators:", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 8)
    for sev, label, detail in indicators:
        color = {"HIGH": (231,76,60), "MEDIUM": (243,156,18), "LOW": (46,204,113)}.get(sev, (0,0,0))
        pdf.set_text_color(*color)
        pdf.set_font("Helvetica", "B", 8)
        pdf.cell(18, 5, f"[{sev}]")
        pdf.set_font("Helvetica", "", 8)
        pdf.set_text_color(0)
        pdf.cell(0, 5, pdf.safe_text(f"{label}: {detail}"), new_x="LMARGIN", new_y="NEXT")
    pdf.ln(4)

    # 2. Operation counts
    pdf.section_title("Operation Category Counts")
    pdf.add_image_safe(p_ops)
    rows_ops = [[cat, int(row["Count"]), f"{row['Pct']}%"] for cat, row in op_counts.iterrows()]
    pdf.add_table(["Category", "Count", "% of Total"], rows_ops, [70, 40, 40])
    pdf.ln(4)

    # 3. Per-site breakdown
    pdf.section_title("Operations by Site (Top 12)")
    if has_site_chart:
        pdf.add_image_safe(p_site)
    cols_site = ["Site"] + list(site_cross.columns)
    n_extra = len(site_cross.columns)
    widths = [55] + [max(12, (190 - 55) / max(n_extra, 1))] * n_extra
    rows_site = []
    for site, row in site_cross.iterrows():
        rows_site.append([str(site)[:30]] + [int(v) for v in row.values])
    pdf.add_table(cols_site, rows_site, widths)
    pdf.ln(4)

    # 4. Daily timeline
    pdf.section_title("Daily Activity Timeline")
    pdf.add_image_safe(p_daily)

    # 5. Hourly distribution
    pdf.section_title("Activity by Hour of Day (UTC)")
    pdf.add_image_safe(p_hour)

    # 6. Last-week vs prior
    pdf.section_title("Last 7 Days vs. Prior Period Comparison")
    lw_cats = lw_df["OpCategory"].value_counts()
    prior_cats = prior_df["OpCategory"].value_counts()
    all_cats = sorted(set(lw_cats.index) | set(prior_cats.index))
    comp_rows = []
    for cat in all_cats:
        lv = int(lw_cats.get(cat, 0))
        pv = int(prior_cats.get(cat, 0))
        prior_days = max((last_week_start - date_min).days, 1)
        pv_daily = pv / prior_days
        lv_daily = lv / 7
        change = ""
        if pv_daily > 0:
            ratio = lv_daily / pv_daily
            change = f"{ratio:.1f}x"
            if ratio > 2:   change += " !!"
            elif ratio > 1.3: change += " !"
        comp_rows.append([cat, pv, f"{pv_daily:.1f}", lv, f"{lv_daily:.1f}", change])
    pdf.add_table(
        ["Category", "Prior Count", "Prior/Day", "Last 7d Count", "Last 7d/Day", "Change"],
        comp_rows, [38, 25, 22, 30, 25, 22])
    pdf.ln(4)

    # 7. Top files in sensitive ops
    pdf.section_title("Top Files in Sensitive Operations (Delete / Download / Share)")
    if not top_files.empty:
        rows_files = [[row["File"][:45], row["Operation"], int(row["Count"])]
                      for _, row in top_files.iterrows()]
        pdf.add_table(["File Name", "Operation", "Count"], rows_files, [100, 45, 25])
    else:
        pdf.set_font("Helvetica", "I", 9)
        pdf.cell(0, 5, "No sensitive file operations found.", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(4)

    # 8. Access context
    pdf.section_title("Access Context: IPs, Platforms, Geolocations")
    ip_counts = udf["ClientIP"].value_counts().head(10)
    if not ip_counts.empty:
        pdf.set_font("Helvetica", "B", 9)
        pdf.cell(0, 5, "Top Client IPs:", new_x="LMARGIN", new_y="NEXT")
        pdf.add_table(["Client IP", "Operations"],
                      [[ip, int(c)] for ip, c in ip_counts.items()], [90, 50])
        pdf.ln(3)

    plat_counts = udf["Platform"].value_counts().head(5)
    if not plat_counts.empty:
        pdf.set_font("Helvetica", "B", 9)
        pdf.cell(0, 5, "Platforms:", new_x="LMARGIN", new_y="NEXT")
        pdf.add_table(["Platform", "Operations"],
                      [[p if p else "N/A", int(c)] for p, c in plat_counts.items()], [90, 50])
        pdf.ln(3)

    geo_counts = udf["GeoLocation"].value_counts().head(5)
    if not geo_counts.empty:
        pdf.set_font("Helvetica", "B", 9)
        pdf.cell(0, 5, "Geolocations:", new_x="LMARGIN", new_y="NEXT")
        pdf.add_table(["Geo", "Operations"],
                      [[g if g else "N/A", int(c)] for g, c in geo_counts.items()], [90, 50])
    pdf.ln(4)

    # 9. Methodology
    pdf.section_title("Methodology")
    pdf.set_font("Helvetica", "", 8)
    pdf.multi_cell(0, 4,
        "This report was automatically generated from Microsoft 365 Unified Audit Log exports. "
        "Operations are categorised from raw audit event names. Risk indicators are heuristic-based "
        "and should be reviewed by qualified personnel. A high risk score does NOT imply wrongdoing - "
        "it highlights activity patterns that warrant further investigation. "
        "All timestamps are in UTC. Off-hours are defined as before 07:00 or after 20:00 UTC.")

    # Save
    safe_name = re.sub(r'[^\w.@-]', '_', user)
    out_path = os.path.join(OUTPUT_DIR, f"audit_report_{safe_name}_{RUN_DATE}.pdf")
    pdf.output(out_path)
    size = os.path.getsize(out_path)
    generated.append((user, out_path, risk_score, len(indicators), size))
    print(f"  -> {out_path}  ({size:,} bytes, risk={risk_score}, {pdf.page_no()} pages)")

# Cleanup
shutil.rmtree(TEMP_DIR, ignore_errors=True)

print(f"\n{'='*60}")
print(f"Done! Generated in '{OUTPUT_DIR}/' on {RUN_DATE}:")
print(f"  {len(excel_generated)} user Excel file(s) (human-readable records)")
print(f"  {len(generated)} PDF report(s)")
for user, path, score, n_ind, size in generated:
    print(f"  {path}  ({size:,} bytes)  risk={score}/100")
